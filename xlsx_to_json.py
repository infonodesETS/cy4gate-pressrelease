import openpyxl
import json
from datetime import date

wb = openpyxl.load_workbook("cy4gate_comunicati_stampa.xlsx")
ws = wb.active

records = []
for row in ws.iter_rows(min_row=2, values_only=True):
    anno, data, ora, titolo, _ = row
    if not titolo:
        continue
    # data è un oggetto date o datetime di openpyxl
    if isinstance(data, date):
        data_str = data.strftime("%d/%m/%Y")
    else:
        data_str = str(data) if data else ""
    records.append({
        "anno": anno,
        "data": data_str,
        "ora": ora or "",
        "titolo": titolo,
        "url": None  # verrà preso dal file xlsx hyperlink
    })

# Rileggi con hyperlinks
wb2 = openpyxl.load_workbook("cy4gate_comunicati_stampa.xlsx")
ws2 = wb2.active
for i, row in enumerate(ws2.iter_rows(min_row=2)):
    if i >= len(records):
        break
    link_cell = row[4]  # colonna E
    if link_cell.hyperlink:
        records[i]["url"] = link_cell.hyperlink.target

with open("comunicati.json", "w", encoding="utf-8") as f:
    json.dump(records, f, ensure_ascii=False, indent=2)

print(f"Salvati {len(records)} record in comunicati.json")
