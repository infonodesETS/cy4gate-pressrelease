import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
import json
import fitz  # PyMuPDF
from datetime import datetime
from pathlib import Path

BASE_URL = "https://www.cy4gate.com"

PAGES = {
    2024: "/company/investor-relations/financial-press-releases/year-2024/",
    2025: "/company/investor-relations/financial-press-releases/year-2025/",
    2026: "/company/investor-relations/financial-press-releases/year-3/",
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}


def parse_date_time(raw: str):
    raw = raw.strip()
    raw = re.sub(r'/+', '/', raw)
    raw = re.sub(r'^(\d{2})\.(\d{2})\.(\d{4})', r'\1/\2/\3', raw)
    parts = raw.split()
    date_str = parts[0] if parts else ""
    time_str = parts[1] if len(parts) > 1 else ""
    try:
        date_obj = datetime.strptime(date_str, "%d/%m/%Y").date()
    except ValueError:
        date_obj = None
    return date_obj, time_str


def download_pdf(url: str, dest_dir: Path) -> Path | None:
    """Scarica un PDF nella cartella dest_dir. Restituisce il path locale o None."""
    filename = url.split("/")[-1].split("?")[0]
    if not filename.endswith(".pdf"):
        filename += ".pdf"
    dest = dest_dir / filename
    if dest.exists():
        return dest  # già scaricato
    try:
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
        if "pdf" not in r.headers.get("Content-Type", "").lower() and not r.content[:4] == b"%PDF":
            return None
        dest.write_bytes(r.content)
        return dest
    except Exception as e:
        print(f"    [WARN] Download fallito {url}: {e}")
        return None


def extract_text(pdf_path: Path) -> str:
    """Estrae il testo da un PDF con PyMuPDF."""
    try:
        doc = fitz.open(pdf_path)
        text = "\n".join(page.get_text() for page in doc)
        doc.close()
        # Pulizia: rimuove righe vuote multiple e strip
        text = re.sub(r'\n{3,}', '\n\n', text).strip()
        return text
    except Exception as e:
        print(f"    [WARN] Estrazione testo fallita {pdf_path}: {e}")
        return ""


def scrape_year(year: int, path: str):
    source_page_url = BASE_URL + path
    print(f"\n{'='*60}")
    print(f"  Scaricando {year}: {source_page_url}")

    # Cartella PDF per questo anno
    pdf_dir = Path(str(year))
    pdf_dir.mkdir(exist_ok=True)

    resp = requests.get(source_page_url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    main = soup.find("div", class_=re.compile(r"main|content|body", re.I))
    if not main:
        main = soup.body

    records = []
    current_date = None
    current_time = None
    seen_pdfs = set()

    elements = main.find_all(["ul", "p", "h1", "h2", "h3"])

    for el in elements:
        if el.name == "ul":
            li_texts = [li.get_text(strip=True) for li in el.find_all("li")]
            for text in li_texts:
                if re.search(r'\d{1,2}[/\.]\d{1,2}[/\.]\d{4}', text):
                    date_obj, time_str = parse_date_time(text)
                    current_date = date_obj
                    current_time = time_str
                    seen_pdfs = set()

        elif el.name == "p" and current_date:
            a = el.find("a", href=True)
            if not a:
                continue
            href = a.get("href", "").strip()
            if not href:
                continue
            if href.startswith("/"):
                full_url = BASE_URL + href
            elif href.startswith("http"):
                full_url = href
            else:
                continue
            if not ("/assets/" in href or href.endswith(".pdf")):
                continue
            if full_url in seen_pdfs:
                continue
            seen_pdfs.add(full_url)

            title = a.get_text(" ", strip=True)
            title = re.sub(r'\s+', ' ', title).strip()

            # Scarica PDF
            print(f"    PDF: {full_url.split('/')[-1]}", end=" ", flush=True)
            pdf_path = download_pdf(full_url, pdf_dir)

            # Estrai testo
            pdf_text = ""
            if pdf_path:
                pdf_text = extract_text(pdf_path)
                print(f"→ {len(pdf_text)} char")
            else:
                print("→ download fallito")

            records.append({
                "anno": year,
                "data": current_date.strftime("%d/%m/%Y") if current_date else "",
                "ora": current_time,
                "titolo": title,
                "url": full_url,
                "source_page": source_page_url,
                "pdf_locale": str(pdf_path) if pdf_path else None,
                "testo": pdf_text,
            })

    print(f"  → {len(records)} comunicati per {year}")
    return records


def build_excel(all_records, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comunicati Stampa"

    headers = ["Anno", "Data", "Ora", "Titolo", "Link", "Source Page"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    link_font = Font(color="0563C1", underline="single")
    alt_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")

    for row_idx, rec in enumerate(all_records, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        ws.cell(row=row_idx, column=1, value=rec["anno"])

        from datetime import date as date_type
        try:
            d = datetime.strptime(rec["data"], "%d/%m/%Y").date()
        except Exception:
            d = rec["data"]
        data_cell = ws.cell(row=row_idx, column=2, value=d)
        data_cell.number_format = "DD/MM/YYYY"

        ws.cell(row=row_idx, column=3, value=rec["ora"])
        ws.cell(row=row_idx, column=4, value=rec["titolo"]).alignment = Alignment(wrap_text=True, vertical="top")

        link_cell = ws.cell(row=row_idx, column=5, value="Apri PDF")
        link_cell.hyperlink = rec["url"]
        link_cell.font = link_font

        sp_cell = ws.cell(row=row_idx, column=6, value=rec["source_page"])
        sp_cell.hyperlink = rec["source_page"]
        sp_cell.font = link_font

        if fill:
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).fill = fill

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 80
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 55
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(all_records) + 1}"

    wb.save(output_path)


if __name__ == "__main__":
    all_records = []
    for year, path in PAGES.items():
        records = scrape_year(year, path)
        all_records.extend(records)

    # Ordina DESC per anno e data
    all_records.sort(
        key=lambda r: (r["anno"], datetime.strptime(r["data"], "%d/%m/%Y") if r["data"] else datetime.min),
        reverse=True
    )

    # Salva JSON (con testo)
    json_path = "comunicati.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_records, f, ensure_ascii=False, indent=2)
    print(f"\nJSON salvato: {json_path} ({len(all_records)} record)")

    # Salva data.js (senza campo testo lungo per il viewer — solo titolo/url/data)
    # Per il viewer include il testo per la ricerca
    with open("data.js", "w", encoding="utf-8") as f:
        f.write("const COMUNICATI = " + json.dumps(all_records, ensure_ascii=False) + ";")
    print(f"data.js salvato")

    # Salva Excel
    xlsx_path = "cy4gate_comunicati_stampa.xlsx"
    build_excel(all_records, xlsx_path)
    print(f"Excel salvato: {xlsx_path}")

    print(f"\nTotale comunicati: {len(all_records)}")
